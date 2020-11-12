#!/usr/bin/env python
# coding: utf-8

# In[ ]:


# przed uruchomieniem skryptu
# należy w bibliotece roboczej posiadać formatkę pliku Zoska.xlsx w celu wygenerowania danych oraz wyżej podane pakiety


# In[1]:


import numpy
import pandas as pd
import os.path
from os import path
import math
from openpyxl import workbook
from openpyxl import load_workbook
import datetime
import random
from faker import Faker
faker = Faker()
import numpy as np


# In[2]:


# z uwagi, że tabele są najbardziej podobne do relacyjnych baz danych będę korzystał z pandas


# In[4]:


def zaczytanie_pliku_zoska(nazwa_pliku):
    # wczytanie pliku zośka
    xls = pd.ExcelFile(nazwa_pliku)
    Polisa = xls.parse(0, header = None)

    dane_polisa = Polisa.iloc[0:13,1:4][Polisa[1].notna()].reset_index(drop=True)

    dane_agenta = Polisa.iloc[14:24,2:4][Polisa[2].notna()]
    dane_agenta = dane_agenta.reset_index(drop=True)


    daty_polisa = Polisa.iloc[0:3,4:6][Polisa[4].notna()]
    daty_polisa = daty_polisa.reset_index(drop=True)


    Polisa = Polisa[Polisa[1].notna()]
    
    # sprawdzanie poprawności danych w zakładce polisa
    if ((dane_polisa.iloc[0,1]  in ('TAK','NIE')) or 
        (not dane_polisa.iloc[2,1].isnull()) or
        (dane_polisa.iloc[2,1] >= 2020) or
        (dane_polisa.iloc[3,1] in range(1,13)) 
        (not dane_polisa.iloc[4,1].isnull()) or
        (not dane_polisa.iloc[6,1].isnull())) :
        if ((dane_polisa.iloc[0,1] == 'TAK') and pd.isnull(dane_polisa.iloc[5,1])) :
            print('Proszę uzupełnić adres')
        else:
            print('Dane w pierwszej części zakładki polisa są poprawne')
    else:
        print('błąd danych ')

    if dane_agenta.iloc[0,1] == 'TAK':
        if (
            dane_agenta.iloc[1,1].isalpha() and
            dane_agenta.iloc[2,1].isalpha() and 
            isinstance(dane_agenta.iloc[3,1], int) and
            isinstance(dane_agenta.iloc[4,1], int) and
            dane_agenta.iloc[5,1] is not None and
            isinstance(dane_agenta.iloc[6,1], int)and
            (isinstance(dane_agenta.iloc[7,1], int) or isinstance(dane_agenta.iloc[7,1], float)) and
            (isinstance(dane_agenta.iloc[8,1], int) or isinstance(dane_agenta.iloc[8,1], float))

        ) :
            print('Dane agenta z zakładki polisa są poprawne')
        else:
            print('Proszę wpisać poprawne dane agenta')
            return()

    if (daty_polisa.iloc[1,1]  is None) or (daty_polisa.iloc[0,1] is None):
        print('Proszę uzupełnić poprawnie daty')
        return()
    else:
        if int(str(daty_polisa.iloc[1,1])[4:6])+1>12:
            data_plus_1 = 1
        else: 
            data_plus_1= int(str(daty_polisa.iloc[1,1])[4:6])+1
        if (
            (str(int(daty_polisa.iloc[0,1]))[-2:] != '01') or
            str(int(daty_polisa.iloc[1,1]))[-2:] != str(datetime.date(int(str(daty_polisa.iloc[1,1])[:4]), data_plus_1, 1) - datetime.timedelta(days=1))[-2:]
        ):
            print('Proszę wpisać poprawnie datę z pierwszym dniem miesiąca i datę z ostatnim dniem miesiąca')
            return()
        else:
            print('Daty zostały poprawnie wpisane')
            
        
    # przekształcam dane z zakładki zgłoszenia na dataframe
    czesc_1 = dane_polisa.T.iloc[:2,:]

    czesc_1 = czesc_1.rename(columns= czesc_1.iloc[0])
    czesc_1 = czesc_1.iloc[1:]
    czesc_1.reset_index(drop=True, inplace=True)

    czesc_2 = daty_polisa.T.iloc[:2,:]
    czesc_2 = czesc_2.rename(columns= czesc_2.iloc[0])
    czesc_2 = czesc_2.iloc[1:]
    czesc_2.reset_index(drop=True, inplace=True)

    czesc_3 = dane_agenta.T.iloc[:2,0:]
    czesc_3 = czesc_3.rename(columns= czesc_3.iloc[0])
    czesc_3 = czesc_3.iloc[1:]
    czesc_3.reset_index(drop=True, inplace=True)

    # tymczasowa tabela zakładki "Polisa"
    tempor_df_polisy = pd.concat([czesc_1, czesc_2, czesc_3], axis=1)
    
    
    # tymczaasowa tabela zakładki "Zgłoszenia"
    insured_tempor = xls.parse(1, usecols = "A:U")
    insured_tempor = insured_tempor[insured_tempor['Unnamed: 0'].notna()]
    insured_tempor = insured_tempor.rename(columns=insured_tempor.iloc[0])
    insured_tempor = insured_tempor.iloc[1:]
    insured_tempor.columns = [ str(col)  for col in insured_tempor.columns[:-3]]+[ 'uposazony_'+str(col)  for col in insured_tempor.columns[-3:]]

    insured_tempor['NrPolisy'] = int(tempor_df_polisy['NrPolisy'].unique())
    
    if tempor_df_polisy['Nowa umowa'][0] == 'TAK':
        tempor_df_polisy['Wpłacona_kwota'] = 0.00
    
    Rezygnacje_temp = xls.parse(2)
    Rezygnacje_temp = Rezygnacje_temp.rename(columns= Rezygnacje_temp.iloc[4])
    Rezygnacje_temp = Rezygnacje_temp.iloc[5:,:]
    
    
    if not path.exists('baza_danych_polisy.json') : 
        # tworzy liste kolumn do bazy polisy i samą bazę polisy
        col_labels = [] 
        col_labels.append(list(dane_polisa.iloc[:,0]))
        col_labels.append(list(daty_polisa.iloc[:,0]))
        col_labels.append(list(dane_agenta.iloc[1:,0]))
        col_labels = [item for sublist in col_labels for item in sublist]
        baza_danych_polisy = pd.DataFrame(columns = col_labels)
        # nowa baza danych to rob nowa kolumne wpłacona kwota:
        baza_danych_polisy['Wpłacona_kwota'] = 0.00
        
    else:
        baza_danych_polisy = pd.read_json('baza_danych_polisy.json')

        # tylko jezeli baza juz istnieje, jezeli nie ma agenta to wyszukuje ostatniego z bazy danych
        if str(tempor_df_polisy['Zmiana'].unique()[0]) == 'NIE':
            for v in list(baza_danych_polisy.columns[9:]):
                tempor_df_polisy[v] = baza_danych_polisy[baza_danych_polisy['NrPolisy'] == int(tempor_df_polisy['NrPolisy'].unique())].iloc[-1, baza_danych_polisy.columns.get_loc(v)]

    # dropuje zmiane, bo nie ma tego w tabeli głównej, po prostu jak jest bez zmiany to bierze ostatnia znaną
    tempor_df_polisy.drop(columns='Zmiana',inplace=True)

    # powiększa baze o nowy plik zośka
    baza_danych_polisy = baza_danych_polisy.append(tempor_df_polisy)

    if not path.exists('baza_danych_ubezp.json'):
        # tworzy bazę danych ubezpieczonych
        Ubezpieczony = xls.parse(1, usecols = "A:U")
        Ubezpieczony = Ubezpieczony[Ubezpieczony['Unnamed: 0'].notna()]
        baza_danych_ubezp = pd.DataFrame(columns=Ubezpieczony.iloc[0])
        baza_danych_ubezp.drop(columns='Zmiana',inplace=True)
        baza_danych_ubezp.columns = baza_danych_ubezp.columns[:-3].union( 'uposazony_'+baza_danych_ubezp.columns[-3:]  )

    else:
        baza_danych_ubezp = pd.read_json('baza_danych_ubezp.json')

        # przypadek braku zmiany uposażonego, bierze ostatnio znanego
        for b in range(len(insured_tempor)):
            if str(insured_tempor['Zmiana'].unique()) == 'NIE':
                for v in list(baza_danych_polisy.columns[-3:0]):
                    insured_tempor.loc[b,v] = baza_danych_ubezp[baza_danych_ubezp['NrPolisy'] == insured_tempor.loc[b,'NrPolisy']].loc[-1,v]

    # tak samo drop kolumny zmiana i powiększenie oryginalnej tablicy

    insured_tempor.drop(columns='Zmiana',inplace=True)
    baza_danych_ubezp = baza_danych_ubezp.append(insured_tempor)

    # jeżeli plik z Rezygnacjami nie istnieje to stwórz nowy 
    if not path.exists('Rezygnacje_main.json'):
        Rezygnacje_main = Rezygnacje_temp
    else:
        Rezygnacje_main = pd.read_json('Rezygnacje_main.json')
        Rezygnacje_main = Rezygnacje_main.append(Rezygnacje_temp)
    
    #baza_danych_ubezp Rezygnacje_temp
    #Jeżeli cos pojawia się w rezygnacjach to dropujemy ten wiersz z tabeli z ubezpieczonymi
    for i in range(len(Rezygnacje_temp)):
        baza_danych_ubezp = baza_danych_ubezp[baza_danych_ubezp['Nr deklaracji'] != Rezygnacje_temp.reset_index().loc[i,'Nr deklaracji']]
    

    Rezygnacje_main.reset_index(drop=True,inplace=True)
    Rezygnacje_main.to_json('Rezygnacje_main.json')
    baza_danych_polisy.reset_index(drop=True,inplace=True)
    baza_danych_polisy.to_json('baza_danych_polisy.json') 
    baza_danych_ubezp.reset_index(drop=True,inplace=True)
    baza_danych_ubezp.to_json('baza_danych_ubezp.json')


# In[5]:


# testowanie/ generowanie listy płatnosci i otwarcie pliku
def generuj_liste_plat(dl):
    lista_polis_bez_pow = []
    baza_danych_polisy = pd.read_json('baza_danych_polisy.json')
    lista_plat = pd.DataFrame(columns = ['Nr konta','Nr polisy','Kwota','Data wpywu','Płatnik'])
    for i in range(dl):
        lista_plat.at[i,'Nr konta'] = random.randint(1000000000000000, 9999999999999999)
        # żeby uniknac powtorek polis na jednej liscie plac
        policy_no = list(baza_danych_polisy['NrPolisy'].unique())[random.randint(0,len(baza_danych_polisy['NrPolisy'].unique()))-1]
        while policy_no in lista_polis_bez_pow:
            policy_no = list(baza_danych_polisy['NrPolisy'].unique())[random.randint(0,len(baza_danych_polisy['NrPolisy'].unique()))-1]
        
        lista_plat.at[i,'Nr polisy'] = policy_no
        lista_polis_bez_pow.append(policy_no)
        
        lista_plat.at[i,'Kwota'] = random.randint(1000, 20000)+random.randint(1, 100)/100
        lista_plat.at[i,'Data wpywu'] = faker.date(pattern='%Y-%m-%d')
        lista_plat.at[i,'Płatnik'] = faker.name()
    lista_plat.to_json('lista_plac.json')


# In[6]:


def przyjecie_listy_plac():
    # odczytuje pliki z bazami
    lista_plat = pd.read_json('lista_plac.json')
    baza_danych_polisy = pd.read_json('baza_danych_polisy.json') 
    baza_danych_ubezp = pd.read_json('baza_danych_ubezp.json')
    baza_danych_polisy['Wpłacona_kwota'] = baza_danych_polisy['Wpłacona_kwota'].astype(float)
    
    # sprawdzaa czy na polisie jest nadpłaata czy nie dopłata 
    for i in range(len(lista_plat)):
        #baza_danych_polisy = pd.read_json('baza_danych_polisy.json') 
        kwota_do_oplacenia = baza_danych_ubezp[baza_danych_ubezp['NrPolisy']== int(lista_plat.loc[i,'Nr polisy'])]['Kwota składki'].sum()
        kwota_wplacona = lista_plat.loc[i,'Kwota']
        roznica =  math.fabs(kwota_wplacona - kwota_do_oplacenia)
        
        # iteruje po bazie polis sprawdzając czy dana polisa jest na liście plac i zwiększa jej warotść, jeżeli jest
        for j in range(len(baza_danych_polisy)):
            if int(baza_danych_polisy.loc[j,'NrPolisy']) == int(lista_plat.loc[i,'Nr polisy']):
                wartosc_prior = float(baza_danych_polisy.loc[j,'Wpłacona_kwota'])
                calosc = wartosc_prior +kwota_wplacona
                baza_danych_polisy.loc[j,'Wpłacona_kwota'] = calosc
        
        
        # printuje czy jest nadpłata, niedopłata czy równo składka
        if kwota_wplacona < kwota_do_oplacenia:
            print('Wpłacona kwota dla polisy nr: '+str(lista_plat.loc[i,'Nr polisy'])+' jest o '+ str(roznica) +'zł za niska na pokrycie wymaganych składek')
        elif kwota_wplacona > kwota_do_oplacenia:
            print('Wpłacona kwota dla polisy nr: '+str(lista_plat.loc[i,'Nr polisy'])+' jest o '+ str(roznica)+'zł za wysoka - jest ponad wymagane składki')
        else:
            print('Wpłacona kwota jest równa wymagalności')

    # zapisuje baze danych z wpłatami
    baza_danych_polisy.to_json('baza_danych_polisy.json')


# In[7]:


polisy_nieopłacone = []
def lista_polis_nieoplaconych():
    j=0
    #wczytuje bazy danych
    baza_danych_polisy = pd.read_json('baza_danych_polisy.json') 
    baza_danych_ubezp = pd.read_json('baza_danych_ubezp.json')
    # sprawdza ile miesięcy minęło od startu polisy do teraz a zatem ile składek miesięcznych powinno wpłynąć
    # porównuje to z kwotą, która wpłynęła ( kolumna wpłacona kwota w bazie polis)
    # tworzy listę polis z niedopłatą
    # jeżeli niedopłata nie jest mniejsza od wartości dwóch składek wtedy usuwa polisę z bazy polis i ubezpieczonych
    
    # ważne należy wpisać dzien miesiąca w którym testujemy, normalnie powinno być 1 i 10 bo tych dni jest generowana lista
    
    if datetime.datetime.today().day == 23 or datetime.datetime.today().day == 10:
        for j in range(len(baza_danych_polisy)):
            
            start_date = baza_danych_polisy.loc[j,'Data zawarcia umowy']
            start_date = datetime.datetime.strptime(str(int(start_date)), '%Y%m%d')
            #start_date = datetime.datetime.strptime(str(int(baza_danych_polisy.loc[j, 'Data zawarcia umowy'])), '%Y%m%d')
            now_date = datetime.datetime.today()
            liczba_mies = (now_date.year - start_date.year) * 12 + (now_date.month - start_date.month)+1
            nalezna_kwota_za_caly_okres = liczba_mies* baza_danych_ubezp[baza_danych_ubezp['NrPolisy']== int(baza_danych_polisy.loc[j,'NrPolisy'])]['Kwota składki'].sum()
            kwota_wplacona_w_okresie =  baza_danych_polisy.loc[j,'Wpłacona_kwota']
            roznica_kwot = kwota_wplacona_w_okresie - nalezna_kwota_za_caly_okres
            if roznica_kwot <0:
                if roznica_kwot>= baza_danych_ubezp[baza_danych_ubezp['NrPolisy']== int(baza_danych_polisy.loc[j,'NrPolisy'])]['Kwota składki'].sum()*2:
                    print('Polisa '+str(baza_danych_polisy.loc[j,'NrPolisy'])+' została usunięta z powodu zbyt dużego zadłużenia.')
                    baza_danych_polisy = baza_danych_polisy[~baza_danych_polisy.NrPolisy.str.contains(str(baza_danych_polisy.loc[j,'NrPolisy']))]
                    baza_danych_ubezp = baza_danych_ubezp[~baza_danych_ubezp.NrPolisy.str.contains(str(baza_danych_polisy.loc[j,'NrPolisy']))]
                else:
                    print('Niedopłata '+str(roznica_kwot)+'zł')
                    polisy_nieopłacone.append(int(baza_danych_polisy.loc[j,'NrPolisy']))
            elif  roznica_kwot>0:
                print('Na polisie istnieje nadpłata w kwocie '+str(roznica_kwot)+ ' zł')
            else:
                print('Składki zostały w pełni opłacone bez nadpłaty')
    print('polisy z niedopłatą to: '+ str(np.unique(polisy_nieopłacone)))


# In[8]:


# dl - ile wierszy powinno być w pliku
# genereuje plik z numerami polis istniejącymi w bazie
def generuj_plik_swiadczen(dl):
    baza_danych_polisy = pd.read_json('baza_danych_polisy.json')
    baza_danych_ubezp = pd.read_json('baza_danych_ubezp.json')
    swiadczenia = pd.DataFrame(columns = ['nr sprawy','Nr polisy','PESEL ubezpieczonego'])
    for i in range(dl):
        swiadczenia.at[i,'nr sprawy'] = random.randint(1,1000)
        # nie bezpośrenie przypisanie poniewać potem będę korzystał ze zmiennej w indeksowaniue a jakbym użył random.randint
        # prawdopodobnie wylosowałby się inny numer
        nr_polisy_ = list(baza_danych_polisy['NrPolisy'].unique())[random.randint(0,len(baza_danych_polisy['NrPolisy'].unique())-1)]
        print(nr_polisy_)
        swiadczenia.at[i,'Nr polisy'] = nr_polisy_
        swiadczenia.at[i,'PESEL ubezpieczonego'] = baza_danych_ubezp[baza_danych_ubezp['NrPolisy'] == nr_polisy_].reset_index().loc[random.randint(0,len(baza_danych_ubezp[baza_danych_ubezp['NrPolisy'] == nr_polisy_])-1),'PESEL']
    
    #zapisuje plik swiadczenia
    swiadczenia.to_json('swiadczenia.json')


# In[9]:


# odczytuje i generuje nowy plik ze świadczeniami
def odczytaj_i_generuj_plik_swiadczen():
    # wczytanie baz danych
    baza_danych_polisy = pd.read_json('baza_danych_polisy.json') 
    baza_danych_ubezp = pd.read_json('baza_danych_ubezp.json')
    swiadczenia = pd.read_json('swiadczenia.json')
    
    # listy z kolumnami adresowymi i dotyczącymi uposażonego
    dane_adresowe = ['Miejscowość','Ulica','Nr domu','nr lokalu','Kod pocztowy','Poczta','Telefon kontaktowy','E-mail']
    uposazony_cols = [col for col in baza_danych_ubezp.columns if 'uposazony' in col]
    
    #przyjmuje plik swiadczenia i sprawdza, czy wpłacona kwota pokrywa wymagalność
    #uzupełnia plik swiadczenia out
    swiadczenia_out = pd.DataFrame(columns=['nr sprawy','czy opłacona','PESEL','Data urodzenia'])
    k=0
    for k in range(len(swiadczenia)):
        # liczba miesiecy -> należna kwota -> różnica między kwotą wpłaconą a należną
        kwota_wplacona_w_calym_okresie =  float(baza_danych_polisy[baza_danych_polisy['NrPolisy'] == swiadczenia.loc[k,'Nr polisy']]['Wpłacona_kwota'].unique())
        start_date = baza_danych_polisy[baza_danych_polisy['NrPolisy'] == swiadczenia.loc[k,'Nr polisy']]['Data zawarcia umowy']
        start_date = datetime.datetime.strptime(str(int(start_date)), '%Y%m%d')
        now_date = datetime.datetime.today()
        liczba_mies = (now_date.year - start_date.year) * 12 + (now_date.month - start_date.month)+1
        
        nalezna_kwota_za_caly_okres = liczba_mies* baza_danych_ubezp[baza_danych_ubezp['NrPolisy']== int(swiadczenia.loc[k,'Nr polisy'])]['Kwota składki'].sum()
        roznica_wplacone_nalezne = kwota_wplacona_w_calym_okresie-nalezna_kwota_za_caly_okres
        # jeżeli różnica jest ujemna to wpisuje NIE, wpisuje też nie jeżeli nie znajduje nr polisy w bazie
        if ((roznica_wplacone_nalezne) <0 or (swiadczenia.loc[k, 'Nr polisy'] not in set(baza_danych_ubezp['NrPolisy'])) or
            (swiadczenia.loc[k, 'PESEL ubezpieczonego'] not in set(baza_danych_ubezp['PESEL']))):
            swiadczenia_out.loc[k, 'czy opłacona'] = 'NIE'
        else:
            swiadczenia_out.loc[k, 'czy opłacona'] = 'TAK'

        # tworzenie tebeli wyjściowej
        swiadczenia_out.at[k,'nr sprawy'] = swiadczenia.loc[k,'nr sprawy']
        swiadczenia_out.at[k,'PESEL'] = swiadczenia.loc[k,'PESEL ubezpieczonego']
        swiadczenia_out.at[k,'Data urodzenia'] = baza_danych_ubezp[(baza_danych_ubezp['NrPolisy'] ==swiadczenia.loc[k,'Nr polisy'])&
                                                          (baza_danych_ubezp['PESEL'] ==swiadczenia.loc[k,'PESEL ubezpieczonego'])].reset_index().loc[0,'Data urodzenia']
        for n in dane_adresowe:
            swiadczenia_out.at[k,n] = list(baza_danych_ubezp[(baza_danych_ubezp['NrPolisy'] ==swiadczenia.loc[k,'Nr polisy'])&
                                                          (baza_danych_ubezp['PESEL'] ==swiadczenia.loc[k,'PESEL ubezpieczonego'])][n])[0]

        for n in uposazony_cols:
            swiadczenia_out.at[k,n] = list(baza_danych_ubezp[(baza_danych_ubezp['NrPolisy'] ==swiadczenia.loc[k,'Nr polisy'])&
                                                          (baza_danych_ubezp['PESEL'] ==swiadczenia.loc[k,'PESEL ubezpieczonego'])][n])[0]
    # zapisanie w pliku 'swiadczenia_out.json'
    swiadczenia_out.to_json('swiadczenia_out.json')


# In[ ]:





# In[ ]:





# In[10]:


#---------------------- TETOWANIE ----------------------------------------------


# In[ ]:





# In[ ]:





# In[11]:


#generowanie nowych umów, tworzy 10 plików Zoska z nowymi polismai i ubezpieczonymi
Nr_deklaracji = 0

for k in range(10):
    wb = load_workbook("Zoska_.xlsx")
    sheets = wb.sheetnames
    Sheet1 = wb[sheets[0]]
    Sheet2 = wb[sheets[1]]
    nowa_umowa = 'TAK'
    nr_polisy = k
    rok = datetime.datetime.today().year
    miesiac = random.randint(1, 12)
    Nazwa = faker.company()
    Adres = faker.address()
    Id_Klienta = random.randint(1, 100000)
    agent_zmiana = 'TAK'
    agent_Nazwisko = faker.last_name()
    agent_imiona = faker.first_name()
    pesel_agent = random.randint(100000, 10000000)
    Tel_kontakt_agent = random.randint(100000000, 999999999)
    email_agent = faker.email()
    nr_konta_agent = random.randint(1000000000000000, 9999999999999999)
    ryczalt_agent  = random.randint(1000, 5000)
    udzial_proc_agent = random.randint(1, 5)/10
    
    # mapowanie numerów wierszy do pól
    d = {
        1:nowa_umowa,
        3:nr_polisy,
        5: rok,
        7: miesiac,
        9: Nazwa,
        11: Adres,
        13:Id_Klienta,
        15: agent_zmiana,
        16: agent_Nazwisko,
        17: agent_imiona,
        18: pesel_agent,
        19: Tel_kontakt_agent,
        20: email_agent,
        21: nr_konta_agent,
        23: ryczalt_agent,
        24: udzial_proc_agent
    }

    for x in d:
        if x <= 13:
            Sheet1.cell(row = x, column = 3).value = d[x]
        else:
            Sheet1.cell(row = x, column = 4).value = d[x]
            
    data_startu = int(faker.date(pattern='%Y%m')+'01')
    Sheet1.cell(row = 1, column = 6).value = data_startu

    data_zakon = 0 
    while data_zakon < data_startu:
        data_zakon = int(faker.date(pattern='%Y%m')+'31')
    Sheet1.cell(row = 2, column = 6).value = data_zakon
    
    
    # zgłoszenia
    zmienna_random = random.randint(2, 100)
    for b in range(1,zmienna_random):
        
        Nr_deklaracji = Nr_deklaracji+1
        Nazwisko = faker.last_name()
        Imie = faker.first_name()
        Dataurodzenia = faker.date(pattern='%Y%m%d')
        Miejsceurodzenia = faker.city()
        PESEL = random.randint(100000, 10000000)
        Obywatelstwo = 'Polskie'
        lista_stosunkow = ['Umowa o prace','Umowa o dzieło','Umowa zlecenie']
        Stosunekprawnypomiędzyubezpieczającymaubezpieczonym = lista_stosunkow[random.randint(0, 2)]
        Kwotaskladki = random.randint(1, 1000)
        Miejscowosc = faker.city()
        Ulica = faker.street_name()
        Nrdomu = random.randint(1, 100)
        nrlokalu = random.randint(1, 10)
        Kodpocztowy = str(random.randint(0, 9))+str(random.randint(0, 9))+'-'+str(random.randint(0, 9))+str(random.randint(0, 9))+str(random.randint(0, 9))
        Poczta = Miejscowosc+' '+Ulica
        Telefonkontaktowy = random.randint(100000000, 999999999)
        Email = faker.email()
        Zmiana = 'NIE'
        Nazwiskoiimie = faker.name()
        Dataimiejsceurodzenia = faker.date()+' '+faker.city()
        Adres = faker.address()
        lista_zmiennych = [
            Nr_deklaracji, Nazwisko, Imie, Dataurodzenia, Miejsceurodzenia, PESEL, Obywatelstwo, 
            Stosunekprawnypomiędzyubezpieczającymaubezpieczonym, Kwotaskladki, Miejscowosc, Ulica, Nrdomu, nrlokalu,
            Kodpocztowy, Poczta, Telefonkontaktowy, Email, Zmiana, Nazwiskoiimie, Dataimiejsceurodzenia, Adres
        ]
        for c in range(21):
            Sheet2.cell(row = b+3, column = c+1).value = lista_zmiennych[c]    
    
    wb.save("Zoska_%d.xlsx"%k) 


# In[12]:


#zaczytuje 5 nowych plików, z nowymi polisami
for i in range(5):
    print('Polisa '+str(i))
    zaczytanie_pliku_zoska("Zoska_%d.xlsx"%i)


# In[13]:


# pokazanie, że polisy zostały załadowane


# In[14]:


baza_danych_polisy = pd.read_json('baza_danych_polisy.json')
baza_danych_polisy


# In[15]:


baza_danych_ubezp = pd.read_json('baza_danych_ubezp.json')
baza_danych_ubezp


# In[16]:


generuj_liste_plat(1)


# In[17]:


lista_plat = pd.read_json('lista_plac.json')
lista_plat


# In[18]:


przyjecie_listy_plac() 


# In[39]:


# w bazie zapisane ile wpłacono kwoty dla danej polisy
baza_danych_polisy = pd.read_json('baza_danych_polisy.json')
baza_danych_polisy


# In[20]:


lista_polis_nieoplaconych()


# In[21]:


generuj_plik_swiadczen(5)


# In[22]:


swiadczenia = pd.read_json('swiadczenia.json')
swiadczenia


# In[23]:


odczytaj_i_generuj_plik_swiadczen()


# In[24]:


swiadczenia_out = pd.read_json('swiadczenia_out.json')
swiadczenia_out


# In[25]:


# rezygnacje
Nr_deklaracji = 0
baza_danych_polisy = pd.read_json('baza_danych_polisy.json') 
baza_danych_ubezp = pd.read_json('baza_danych_ubezp.json')

for k in range(len(baza_danych_polisy)):
    # wczytanie oryginalnego pustego pliku
    wb = load_workbook("Zoska_.xlsx")
    sheets = wb.sheetnames
    Sheet1 = wb[sheets[0]]
    Sheet3 = wb[sheets[2]]
    # wczytanie zakładek

    #uzupełnianie
    nowa_umowa = 'NIE'
    nr_polisy = baza_danych_polisy.loc[k,'NrPolisy']
    rok = datetime.datetime.today().year
    miesiac = random.randint(1, 12)
    Nazwa = baza_danych_polisy.loc[k,'Nazwa ']
    Adres = baza_danych_polisy.loc[k,'Adres']
    Id_Klienta = baza_danych_polisy.loc[k,'IdKlienta']
    agent_zmiana = 'NIE'

    
    # mapowanie numerów wierszy do pól
    d = {
        1:nowa_umowa,
        3:nr_polisy,
        5: rok,
        7: miesiac,
        9: Nazwa,
        11: Adres,
        13:Id_Klienta,
        15: agent_zmiana,

    }

    for x in d:
        if x <= 13:
            Sheet1.cell(row = x, column = 3).value = d[x]
        else:
            Sheet1.cell(row = x, column = 4).value = d[x]
            
    data_startu = baza_danych_polisy.loc[k,'Data zawarcia umowy']
    Sheet1.cell(row = 1, column = 6).value = data_startu

    data_zakon = baza_danych_polisy.loc[k,'Data ważności umowy']
    Sheet1.cell(row = 2, column = 6).value = data_zakon#str(2020)+str(random.randint(1, 12))+str('31')
    
    
    baza_ubezp = pd.read_json('baza_danych_ubezp.json')
    for m in range(1,random.randint(2,10)):
        if len(baza_danych_ubezp) != 0:
            losowy_rezygnujacy = random.randint(0,len(list(baza_ubezp['Nazwisko'])))
            Nr_deklaracji = list(baza_ubezp['Nr deklaracji'])[losowy_rezygnujacy]
            Nazwisko = list(baza_ubezp['Nazwisko'])[losowy_rezygnujacy]
            Imie = list(baza_ubezp['Imię'])[losowy_rezygnujacy]
            Dataurodzenia =  list(baza_ubezp['Data urodzenia'])[losowy_rezygnujacy]
            PESEL = list(baza_ubezp['Data urodzenia'])[losowy_rezygnujacy]
            lista_powod = ['Zaakonczenie współpracy','rezygnacja']
            Powod = lista_powod[random.randint(0,1)]

            lista_kol_rez = [Nr_deklaracji,Nazwisko,Imie,Dataurodzenia,PESEL, Powod]
            for l in range(0,6):
                Sheet3.cell(row = m+6, column = l+1).value = lista_kol_rez[l]

    wb.save("Zoska_r_%d.xlsx"%k) 


# In[26]:


for i in range(len(pd.read_json('baza_danych_polisy.json'))):
    zaczytanie_pliku_zoska("Zoska_r_%d.xlsx"%i)


# In[27]:


Rezygnacje_main = pd.read_json('Rezygnacje_main.json')
Rezygnacje_main


# In[28]:


#Pokazuje, że ubezp z zakładki rezygnacje zostały skasowane z bazy danych ubezpieczonych
baza_danych_ubezp = pd.read_json('baza_danych_ubezp.json')
for i in list(Rezygnacje_main['Nr deklaracji']):
    print(baza_danych_ubezp[baza_danych_ubezp['Nr deklaracji']==i])


# In[29]:


# jak widać ubezpieczeni zostali usunięci


# In[30]:


#istniejące - zmiana dodanie ubezpieczonych


# In[31]:


Nr_deklaracji = max(baza_danych_ubezp['Nr deklaracji'])


# In[32]:


Nr_deklaracji


# In[33]:


Nr_deklaracji = max(baza_danych_ubezp['Nr deklaracji'])
baza_danych_polisy = pd.read_json('baza_danych_polisy.json') 
baza_danych_ubezp = pd.read_json('baza_danych_ubezp.json')
for k in range(10):
    ran_num = random.randint(0,len(baza_danych_polisy)-1)
    
    wb = load_workbook("Zoska_.xlsx")
    sheets = wb.sheetnames
    Sheet1 = wb[sheets[0]]
    Sheet2 = wb[sheets[1]]
    nowa_umowa = 'NIE'
    nr_polisy = baza_danych_polisy.loc[ran_num,'NrPolisy']
    rok = datetime.datetime.today().year
    miesiac = random.randint(1, 12)
    Nazwa = baza_danych_polisy.loc[ran_num,'Nazwa ']
    Adres = baza_danych_polisy.loc[ran_num,'Adres']
    Id_Klienta = baza_danych_polisy.loc[ran_num,'IdKlienta']
    agent_zmiana = 'NIE'
    
    # mapowanie numerów wierszy do pól
    d = {
        1:nowa_umowa,
        3:nr_polisy,
        5: rok,
        7: miesiac,
        9: Nazwa,
        11: Adres,
        13:Id_Klienta,
        15: agent_zmiana
        
    }

    for x in d:
        if x <= 13:
            Sheet1.cell(row = x, column = 3).value = d[x]
        else:
            Sheet1.cell(row = x, column = 4).value = d[x]
            
    data_startu = baza_danych_polisy.loc[ran_num,'Data zawarcia umowy']
    Sheet1.cell(row = 1, column = 6).value = data_startu

    data_zakon =  baza_danych_polisy.loc[ran_num,'Data ważności umowy']
    Sheet1.cell(row = 2, column = 6).value = data_zakon
    
    
    # zgłoszenia
    # losowy wybór ilości ubezpieczoych
    zmienna_random = random.randint(1, 100)
    
    # dowiaduje się jaki jest max numer deklaracji w istniejącej bazie danych
    
    
    for b in range(1,zmienna_random):
        
        # symulowanie danych ubezpieczonych
        Nr_deklaracji = Nr_deklaracji +1
        
        Nazwisko = faker.last_name()
        Imie = faker.first_name()
        Dataurodzenia = faker.date(pattern='%Y%m%d')
        Miejsceurodzenia = faker.city()
        PESEL = random.randint(100000, 10000000)
        Obywatelstwo = 'Polskie'
        lista_stosunkow = ['Umowa o prace','Umowa o dzieło','Umowa zlecenie']
        Stosunekprawnypomiędzyubezpieczającymaubezpieczonym = lista_stosunkow[random.randint(0, 2)]
        Kwotaskladki = random.randint(1, 1000)
        Miejscowosc = faker.city()
        Ulica = faker.street_name()
        Nrdomu = random.randint(1, 100)
        nrlokalu = random.randint(1, 10)
        Kodpocztowy = str(random.randint(0, 9))+str(random.randint(0, 9))+'-'+str(random.randint(0, 9))+str(random.randint(0, 9))+str(random.randint(0, 9))
        Poczta = Miejscowosc+' '+Ulica
        Telefonkontaktowy = random.randint(100000000, 999999999)
        Email = faker.email()
        Zmiana = 'TAK'
        Nazwiskoiimie = faker.name()
        Dataimiejsceurodzenia = faker.date()+' '+faker.city()
        Adres = faker.address()
        lista_zmiennych = [
            Nr_deklaracji, Nazwisko, Imie, Dataurodzenia, Miejsceurodzenia, PESEL, Obywatelstwo, 
            Stosunekprawnypomiędzyubezpieczającymaubezpieczonym, Kwotaskladki, Miejscowosc, Ulica, Nrdomu, nrlokalu,
            Kodpocztowy, Poczta, Telefonkontaktowy, Email, Zmiana, Nazwiskoiimie, Dataimiejsceurodzenia, Adres
        ]
        for c in range(21):
            Sheet2.cell(row = b+3, column = c+1).value = lista_zmiennych[c]    
    
    #zapisanie w pliku
    wb.save("Zoska_ub_%d.xlsx"%k) 


# In[34]:


baza_danych_ubezp = pd.read_json('baza_danych_ubezp.json')


# In[35]:


for i in range(4):
    zaczytanie_pliku_zoska("Zoska_ub_%d.xlsx"%i)


# In[36]:


# baza polis powiększa się ponieważ rok i miesiac zmian na polissie może się zmieniać


# In[37]:


baza_danych_polisy = pd.read_json('baza_danych_polisy.json')
baza_danych_polisy


# In[38]:


# po dodaniu polis, zwiększa się kwota tylko w konkrtenych wierszach, dla której polisy mamy nową wpłatę
przyjecie_listy_plac() 


# In[ ]:




